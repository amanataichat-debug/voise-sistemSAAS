"""
Contact Import Service — массовый импорт контактов агента из xlsx/csv.

Возможности:
- Парсинг .xlsx (openpyxl) и .csv (UTF-8, разделители ';' и ',').
- Детект заголовков по словарю синонимов; fallback — парсинг по индексу колонки.
- Нормализация и валидация телефонов (phonenumbers).
- Распределение времени звонков параллельно по 5 (группы по 30 секунд).
- Проверка рабочих часов агента (МСК) с переносом на следующий рабочий день.
- Временное хранилище распарсенного превью (JSON-файл с TTL 30 минут).
- Генерация шаблона xlsx и файла с ошибками.

Время: ввод из xlsx/csv трактуется как МСК, в БД пишется UTC.
"""

import io
import os
import re
import csv
import json
import uuid
import time
import tempfile
from datetime import datetime, timedelta, timezone
from typing import Optional, List, Dict, Any, Tuple

import phonenumbers
from openpyxl import Workbook, load_workbook

from backend.core.logging import get_logger
from backend.core.timezone_utils import (
    msk_to_utc,
    adjust_to_working_hours,
    now_utc,
)

logger = get_logger(__name__)

# ============================================================================
# КОНСТАНТЫ
# ============================================================================

MAX_IMPORT_ROWS = 1000
# Консервативная оценка стоимости одного контакта (PreCall + PostCall) в кредитах.
CREDITS_PER_CONTACT = 30

PREVIEW_TTL_SECONDS = 30 * 60  # 30 минут

# Распределение времени звонков: группы по 5, шаг 30 секунд, старт через 60 минут.
PARALLEL_GROUP_SIZE = 5
GROUP_INTERVAL_SECONDS = 30
FIRST_GROUP_DELAY_MINUTES = 60

# Временное хранилище превью
_PREVIEW_DIR = os.path.join(tempfile.gettempdir(), "voicyfy_contact_imports")

# Порядок колонок при парсинге по индексу (нет заголовков)
COLUMN_ORDER = [
    "name",            # A
    "phone",           # B
    "company",         # C
    "position",        # D
    "notes",           # E
    "task_title",      # F
    "task_description",  # G
    "scheduled_time",  # H
]

# Синонимы заголовков (lowercase, без лишних пробелов)
HEADER_SYNONYMS = {
    "name": {"имя", "name", "фио", "имя контакта", "контакт"},
    "phone": {"телефон", "phone", "номер", "тел", "номер телефона", "mobile", "моб"},
    "company": {"компания", "company", "организация", "фирма"},
    "position": {"должность", "position", "роль", "title"},
    "notes": {"информация о клиенте", "заметки", "notes", "комментарий",
              "инфо", "note", "описание клиента"},
    "task_title": {"задача", "задача (заголовок)", "title", "task", "заголовок задачи"},
    "task_description": {"описание задачи", "описание", "description"},
    "scheduled_time": {"когда звонить", "время", "scheduled_time", "дата",
                       "когда", "время звонка", "дата и время"},
}

# Форматы дат для csv / текстовых значений (xlsx datetime приходит как объект)
_DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M",
    "%d.%m.%Y %H:%M:%S",
    "%d.%m.%Y %H:%M",
    "%d.%m.%Y",
    "%Y-%m-%d",
    "%d/%m/%Y %H:%M",
    "%d/%m/%Y",
]


# ============================================================================
# НОРМАЛИЗАЦИЯ ТЕЛЕФОНОВ
# ============================================================================

def normalize_phone(raw: Any) -> Tuple[Optional[str], Optional[str]]:
    """
    Нормализует и валидирует телефон.
    Возвращает (E164 номер, None) при успехе либо (None, текст ошибки).
    """
    if raw is None:
        return None, "Телефон обязателен"

    # openpyxl может вернуть число (79001234567) как int/float
    if isinstance(raw, float):
        raw = str(int(raw)) if raw == int(raw) else str(raw)
    elif isinstance(raw, int):
        raw = str(raw)

    s = str(raw).strip()
    if not s:
        return None, "Телефон обязателен"

    # Убрать пробелы, скобки, дефисы, точки
    cleaned = re.sub(r"[\s\(\)\-\.]", "", s)
    if not cleaned:
        return None, "Телефон обязателен"

    if cleaned.startswith("+"):
        candidate = cleaned
    elif cleaned.startswith("8") and len(cleaned) == 11:
        candidate = "+7" + cleaned[1:]
    elif cleaned.startswith("7"):
        candidate = "+" + cleaned
    else:
        # Остальное пытаемся распарсить с регионом RU
        candidate = cleaned

    try:
        num = phonenumbers.parse(candidate, "RU")
        if not phonenumbers.is_valid_number(num):
            return None, f"Невалидный телефон: '{s}'"
        return phonenumbers.format_number(
            num, phonenumbers.PhoneNumberFormat.E164
        ), None
    except phonenumbers.NumberParseException:
        return None, f"Невалидный телефон: '{s}'"


# ============================================================================
# ПАРСИНГ ДАТЫ (МСК)
# ============================================================================

def _parse_datetime_msk(raw: Any) -> Tuple[Optional[datetime], Optional[str]]:
    """
    Парсит значение времени (трактуется как МСК) → naive datetime в МСК.
    Возвращает (datetime|None, error|None). Пустое значение → (None, None).
    """
    if raw is None:
        return None, None

    if isinstance(raw, datetime):
        # openpyxl вернул datetime — трактуем как МСК (naive)
        return raw.replace(tzinfo=None), None

    s = str(raw).strip()
    if not s:
        return None, None

    # ISO с маркером таймзоны
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is not None:
            # Уже aware — оставляем как есть (нормализуется позже)
            return dt, None
        return dt, None
    except ValueError:
        pass

    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt), None
        except ValueError:
            continue

    return None, f"Не распознано время: '{s}'"


# ============================================================================
# ДЕТЕКТ ЗАГОЛОВКОВ
# ============================================================================

def _norm_header(val: Any) -> str:
    return re.sub(r"\s+", " ", str(val or "").strip().lower())


def detect_headers(first_row: List[Any]) -> Optional[Dict[int, str]]:
    """
    Если первая строка содержит >= 2 узнаваемых заголовка — возвращает
    маппинг {col_index: field_name}. Иначе None (парсим по индексу).
    """
    mapping: Dict[int, str] = {}
    for idx, cell in enumerate(first_row):
        h = _norm_header(cell)
        if not h:
            continue
        for field, synonyms in HEADER_SYNONYMS.items():
            if h in synonyms:
                mapping[idx] = field
                break

    if len(mapping) >= 2:
        return mapping
    return None


# ============================================================================
# ЧТЕНИЕ ФАЙЛА В МАТРИЦУ СТРОК
# ============================================================================

def _read_rows(filename: str, content: bytes) -> List[List[Any]]:
    """Читает xlsx/csv в список строк (каждая — список ячеек)."""
    name = (filename or "").lower()

    if name.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        wb.close()
        return rows

    if name.endswith(".csv"):
        text = None
        for enc in ("utf-8-sig", "utf-8", "cp1251"):
            try:
                text = content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            text = content.decode("utf-8", errors="replace")

        # Определяем разделитель (';' или ',')
        sample = text[:4096]
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        return [list(r) for r in reader]

    raise ValueError("unsupported_format")


# ============================================================================
# ПАРСИНГ ФАЙЛА
# ============================================================================

def parse_file(filename: str, content: bytes) -> Dict[str, Any]:
    """
    Парсит файл в структуру:
    {
      "rows": [ {row, name, phone, company, position, notes,
                 task_title, task_description, scheduled_dt(naive msk|None|aware)} ],
      "errors": [ {row, error} ],
      "total_rows": N,
    }
    rows — только валидные строки (с корректным телефоном).
    """
    raw_rows = _read_rows(filename, content)

    # Убираем полностью пустые строки в конце
    while raw_rows and all(
        (c is None or str(c).strip() == "") for c in raw_rows[-1]
    ):
        raw_rows.pop()

    if not raw_rows:
        return {"rows": [], "errors": [], "total_rows": 0}

    header_map = detect_headers(raw_rows[0])
    if header_map is not None:
        data_rows = raw_rows[1:]
        start_offset = 2  # человекочитаемый номер строки (1 — заголовок)
        col_map = header_map
    else:
        data_rows = raw_rows
        start_offset = 1
        col_map = {i: COLUMN_ORDER[i] for i in range(len(COLUMN_ORDER))}

    rows: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    total = 0

    for i, raw in enumerate(data_rows):
        row_num = start_offset + i

        # Пропускаем пустые строки
        if not raw or all((c is None or str(c).strip() == "") for c in raw):
            continue

        total += 1

        def cell(field: str) -> Any:
            for idx, fname in col_map.items():
                if fname == field and idx < len(raw):
                    return raw[idx]
            return None

        phone_raw = cell("phone")
        phone, perr = normalize_phone(phone_raw)
        if perr:
            errors.append({"row": row_num, "error": perr})
            continue

        scheduled_dt, derr = _parse_datetime_msk(cell("scheduled_time"))
        if derr:
            errors.append({"row": row_num, "error": derr})
            continue

        def s(field: str) -> Optional[str]:
            v = cell(field)
            if v is None:
                return None
            v = str(v).strip()
            return v or None

        rows.append({
            "row": row_num,
            "name": s("name"),
            "phone": phone,
            "company": s("company"),
            "position": s("position"),
            "notes": s("notes"),
            "task_title": s("task_title"),
            "task_description": s("task_description"),
            "scheduled_dt": scheduled_dt,  # naive(МСК) | aware | None
        })

    return {"rows": rows, "errors": errors, "total_rows": total}


# ============================================================================
# РАСПРЕДЕЛЕНИЕ ВРЕМЕНИ + РАБОЧИЕ ЧАСЫ
# ============================================================================

def _default_scheduled_utc(group_index: int, base_utc: datetime) -> datetime:
    """
    Дефолтное время для группы (по 5 контактов):
    группа 0 → base + 60 минут, далее +30 сек на каждую следующую группу.
    """
    return (
        base_utc
        + timedelta(minutes=FIRST_GROUP_DELAY_MINUTES)
        + timedelta(seconds=GROUP_INTERVAL_SECONDS * group_index)
    )


def assign_schedule(
    rows: List[Dict[str, Any]],
    wh_start: int,
    wh_end: int,
    base_utc: Optional[datetime] = None,
) -> int:
    """
    Проставляет каждому row 'scheduled_time_utc' (ISO-строка) и 'shifted' (bool).

    - Если в строке указано время — конвертируем МСК → UTC.
    - Иначе — дефолтное распределение по группам.
    - Затем приводим к рабочим часам агента (МСК), считаем перенесённые.

    Возвращает количество строк, сдвинутых из-за рабочих часов.
    """
    if base_utc is None:
        base_utc = now_utc()

    shifted_count = 0

    for idx, row in enumerate(rows):
        dt = row.get("scheduled_dt")
        if dt is not None:
            # Явно указанное время: naive → МСК; aware → как есть в UTC
            if dt.tzinfo is None:
                target_utc = msk_to_utc(dt)
            else:
                target_utc = dt.astimezone(timezone.utc)
        else:
            group_index = idx // PARALLEL_GROUP_SIZE
            target_utc = _default_scheduled_utc(group_index, base_utc)

        adjusted_utc, shifted = adjust_to_working_hours(target_utc, wh_start, wh_end)
        if shifted:
            shifted_count += 1

        row["scheduled_time_utc"] = adjusted_utc.isoformat()
        row["shifted"] = shifted

    return shifted_count


# ============================================================================
# ВРЕМЕННОЕ ХРАНИЛИЩЕ ПРЕВЬЮ
# ============================================================================

def _ensure_preview_dir():
    os.makedirs(_PREVIEW_DIR, exist_ok=True)


def _preview_path(token: str) -> str:
    # token — UUID, защищаемся от path traversal
    safe = re.sub(r"[^a-zA-Z0-9\-]", "", token)
    return os.path.join(_PREVIEW_DIR, f"{safe}.json")


def save_preview(data: Dict[str, Any]) -> str:
    """Сохраняет распарсенное превью во временный файл, возвращает preview_token."""
    _ensure_preview_dir()
    token = str(uuid.uuid4())
    data = dict(data)
    data["_created_at"] = time.time()
    with open(_preview_path(token), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, default=str)
    _cleanup_expired()
    return token


def load_preview(token: str) -> Optional[Dict[str, Any]]:
    """Загружает превью по токену. None если не найдено или истёк TTL."""
    path = _preview_path(token)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError):
        return None

    created = data.get("_created_at", 0)
    if time.time() - created > PREVIEW_TTL_SECONDS:
        try:
            os.remove(path)
        except OSError:
            pass
        return None
    return data


def delete_preview(token: str):
    try:
        os.remove(_preview_path(token))
    except OSError:
        pass


def _cleanup_expired():
    """Удаляет просроченные превью-файлы."""
    try:
        for fname in os.listdir(_PREVIEW_DIR):
            if not fname.endswith(".json"):
                continue
            fpath = os.path.join(_PREVIEW_DIR, fname)
            try:
                if time.time() - os.path.getmtime(fpath) > PREVIEW_TTL_SECONDS:
                    os.remove(fpath)
            except OSError:
                continue
    except OSError:
        pass


# ============================================================================
# ГЕНЕРАЦИЯ XLSX (ШАБЛОН / ОШИБКИ)
# ============================================================================

def generate_template_xlsx() -> bytes:
    """Генерирует xlsx-шаблон с заголовками и одной строкой-примером."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Контакты"

    headers = [
        "Имя", "Телефон", "Компания", "Должность",
        "Информация о клиенте", "Задача (заголовок)",
        "Описание задачи", "Когда звонить (МСК)",
    ]
    ws.append(headers)

    example = [
        "Иван Петров", "+79001234567", "ООО Ромашка", "Директор",
        "Интересовался тарифом Профи, просил перезвонить",
        "Презентация продукта", "Рассказать о новых возможностях",
        "10.06.2026 14:30",
    ]
    ws.append(example)

    # Ширина колонок для читаемости
    widths = [18, 18, 18, 16, 32, 24, 28, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_errors_xlsx(errors: List[Dict[str, Any]]) -> bytes:
    """Генерирует xlsx с проблемными строками (номер строки + причина)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ошибки"
    ws.append(["Строка", "Ошибка"])
    for e in errors:
        ws.append([e.get("row"), e.get("error")])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 60
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
